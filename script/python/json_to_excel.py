import pandas as pd
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def json_to_excel_advanced(
    json_path: str,
    excel_path: str = None,
    table_style: str = "Medium",          # Light / Medium / None
    multiple_sheets: bool = False,        # 是否把多个顶级数组放到不同工作表
    convert_numeric: bool = True
):
    """
    更接近 Aspose 在线工具的 JSON → Excel 转换
    支持：
      - 自动展开嵌套对象
      - 可选表格样式
      - 多工作表
    """
    json_path = Path(json_path)
    if excel_path is None:
        excel_path = json_path.with_suffix(".xlsx")

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    # 删除默认的空 sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    def apply_table_style(ws, style="Medium"):
        if style == "None":
            return
        header_fill = PatternFill(
            start_color="4472C4" if style == "Medium" else "D9E2F3",
            end_color="4472C4" if style == "Medium" else "D9E2F3",
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF" if style == "Medium" else "000000")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border

    def write_df_to_sheet(df, sheet_name):
        if convert_numeric:
            for col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="ignore")
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel 工作表名最长 31 字符
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        apply_table_style(ws, table_style)
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # 处理不同 JSON 结构
    if isinstance(data, list):
        df = pd.json_normalize(data)
        write_df_to_sheet(df, "Sheet1")
    elif isinstance(data, dict):
        if multiple_sheets:
            for key, value in data.items():
                if isinstance(value, list):
                    df = pd.json_normalize(value)
                    write_df_to_sheet(df, str(key))
                else:
                    # 非数组的值单独放一个 sheet
                    df = pd.json_normalize({key: value})
                    write_df_to_sheet(df, str(key))
        else:
            # 默认把所有内容压平到一个工作表
            df = pd.json_normalize(data)
            write_df_to_sheet(df, "Sheet1")
    else:
        raise ValueError("不支持的 JSON 结构")

    wb.save(excel_path)
    print(f"转换成功 → {excel_path}")
    return excel_path


# -------------------- 使用示例 --------------------
if __name__ == "__main__":
    # 基础用法
    json_to_excel_advanced("data.json", "output.xlsx")

    # 带样式 + 多工作表
    # json_to_excel_advanced(
    #     "data.json",
    #     "output_styled.xlsx",
    #     table_style="Medium",
    #     multiple_sheets=True
    # )